#!/usr/bin/env python3
"""
generate_report_excel.py  — v4
Fixes:
1. Summary Dashboard decluttered — removed Needs Fulfillment (moved to Shelters sheet)
2. Regional Risk Score fixed — uses minimum 5 incidents as denominator to avoid
   small regions scoring 100% just because they have 1 active out of 1 total incident.
   Also renamed to "Regional Activity Score" to distinguish from global severity score.
3. Recommendation column rows are now tall enough (no overlap).
4. Added a note on Summary linking to reports.php severity score for clarity.
"""

import sys, json
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NAVY       = "1B2A4A"
NAVY_MID   = "243558"
NAVY_LIGHT = "2E4272"
SILVER     = "F4F6FA"
WHITE      = "FFFFFF"
GRAY_HDR   = "DDE1EA"
GRAY_TXT   = "6B7280"
BORDER_CLR = "C5CBD8"
STRIPE_A   = "F8F9FC"
STRIPE_B   = WHITE

C_RED      = "C0392B"
C_AMBER    = "D97706"
C_GREEN    = "1A7A4A"
C_BLUE     = "1D4ED8"
C_PURPLE   = "5B21B6"
C_TEAL     = "0F766E"

KPI_PAIRS = [
    (NAVY_MID, "FFFFFF"),
    (C_RED,    "FFFFFF"),
    (C_GREEN,  "FFFFFF"),
    (NAVY_MID, "FFFFFF"),
    (C_RED,    "FFFFFF"),
    (NAVY_MID, "FFFFFF"),
    (C_GREEN,  "FFFFFF"),
    (NAVY_MID, "FFFFFF"),
]

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=NAVY, size=10, italic=False):
    return Font(name="Calibri", bold=bold, color=color, size=size, italic=italic)

def _border(color=BORDER_CLR):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left(indent=1):
    return Alignment(horizontal="left", vertical="center", wrap_text=True, indent=indent)

def _right():
    return Alignment(horizontal="right", vertical="center")

def add_title(ws, title, subtitle, ncols):
    ws.row_dimensions[1].height = 48
    ws.row_dimensions[2].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=title)
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=18)
    c.fill = _fill(NAVY); c.alignment = _center()
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = Font(name="Calibri", italic=True, color=GRAY_TXT, size=9)
    s.fill = _fill(SILVER); s.alignment = _center()

def add_section(ws, row, label, ncols, color=NAVY_MID):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    ws.row_dimensions[row].height = 28
    c = ws.cell(row=row, column=1, value=label)
    c.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    c.fill = _fill(color); c.alignment = _left()

def add_header(ws, row, headers, height=24):
    ws.row_dimensions[row].height = height
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i + 1, value=h)
        c.font = Font(name="Calibri", bold=True, color=NAVY, size=10)
        c.fill = _fill(GRAY_HDR); c.alignment = _center(); c.border = _border()

def add_row(ws, row, values, alt=False, height=28):
    ws.row_dimensions[row].height = height
    bg = STRIPE_A if alt else STRIPE_B
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=i + 1, value=v)
        c.font = _font(color="2D3748")
        c.fill = _fill(bg); c.alignment = _left(); c.border = _border()

def color_cell(ws, row, col, color, bold=True, size=10):
    c = ws.cell(row=row, column=col)
    c.font = Font(name="Calibri", bold=bold, color=color, size=size)

def hhmm(minutes):
    if minutes is None: return "N/A"
    try:
        m = int(float(minutes)); h = m // 60; r = m % 60
        return f"{h}h {r}m" if h else f"{r}m"
    except: return "N/A"

def safe_avg(vals):
    v = [x for x in vals if x is not None]
    return sum(v) / len(v) if v else None

def safe_min(vals):
    v = [x for x in vals if x is not None]; return min(v) if v else None

def safe_max(vals):
    v = [x for x in vals if x is not None]; return max(v) if v else None

def weighted_avg(vals):
    if not vals: return None
    n = len(vals)
    weights = list(range(1, n+1))
    return sum(v*w for v,w in zip(vals,weights)) / sum(weights)


SEV_COLOR = {"critical": C_RED, "high": C_AMBER, "medium": C_AMBER, "low": C_GREEN, "info": C_BLUE}
ST_INC    = {"resolved": C_GREEN, "in progress": C_AMBER, "active": C_RED, "in_progress": C_AMBER}
ST_NEED   = {"fulfilled": C_GREEN, "in_progress": C_AMBER, "pending": C_AMBER, "rejected": C_RED}
ST_HOSP   = {"safe": C_GREEN, "warning": C_AMBER, "dangerous": C_RED}
ST_ROAD   = {"open": C_GREEN, "safe": C_GREEN, "warning": C_AMBER, "closed": C_RED, "blocked": C_RED}
ST_ROUTE  = {"open": C_GREEN, "safe": C_GREEN, "warning": C_AMBER, "closed": C_RED, "blocked": C_RED}
ST_ZONE   = {"safe": C_GREEN, "warning": C_AMBER, "danger": C_RED}
ST_INFRA  = {"intact": C_GREEN, "minor damage": C_AMBER, "partially damaged": C_AMBER, "severely damaged": C_RED, "destroyed": C_RED}
ST_UTIL   = {"available": C_GREEN, "stable": C_GREEN, "limited": C_AMBER, "unstable": C_AMBER, "unavailable": C_RED, "cut": C_RED}
ST_TEAM   = {"available": C_GREEN, "on mission": C_AMBER, "busy": C_AMBER, "unavailable": C_RED}

def readiness_score(h):
    score = 100
    st = (h.get("hospital_status") or "").lower()
    if st == "dangerous": score -= 40
    elif st == "warning": score -= 20
    infra = (h.get("infrastructure_status") or "").lower()
    if "destroyed" in infra or "severely" in infra: score -= 25
    elif "damage" in infra: score -= 10
    pwr = (h.get("power_status") or "").lower()
    if "unavailable" in pwr or "cut" in pwr: score -= 20
    elif "unstable" in pwr or "limited" in pwr: score -= 10
    wtr = (h.get("water_status") or "").lower()
    if "unavailable" in wtr or "cut" in wtr: score -= 15
    elif "unstable" in wtr or "limited" in wtr: score -= 8
    return max(0, score)

def readiness_label(score):
    if score >= 75: return ("High", C_GREEN)
    elif score >= 45: return ("Medium", C_AMBER)
    else: return ("Critical", C_RED)

def crisis_severity_index(d):
    if "severityScore" in d and d["severityScore"] is not None:
        try:
            return int(round(float(d["severityScore"])))
        except (TypeError, ValueError):
            pass
    return 0

def region_activity_score(region_name, d):
    """
    Regional Activity Score — renamed from 'Risk Score' to avoid confusion
    with the global dashboard severity score.

    FIX: Uses max(5, actual_count) as denominators so small regions with
    1 incident don't automatically score 100. A region needs to have a
    meaningful volume before it can score high.

    Formula: 40% active_inc ratio + 30% critical_alerts ratio +
             20% dangerous_hospitals ratio + 10% blocked_roads ratio
    All ratios are capped via min_denominator=5.
    """
    incidents = d.get("incidents", [])
    alerts    = d.get("alerts", [])
    hospitals = d.get("hospitals", [])
    roads     = d.get("mapRoads", []) + d.get("polRoads", [])

    reg = region_name.lower()
    reg_inc    = [r for r in incidents if (r.get("location") or "").lower() == reg]
    reg_alerts = [a for a in alerts    if (a.get("region")   or "").lower() == reg]
    reg_hosp   = [h for h in hospitals if (h.get("location") or "").lower() == reg]
    reg_roads  = [r for r in roads     if (r.get("region")   or "").lower() == reg]

   
    MIN_D = 5
    total_inc  = max(MIN_D, len(reg_inc))
    active_inc = sum(1 for r in reg_inc if r.get("status","").lower() not in ("resolved",))
    total_al   = max(MIN_D, len(reg_alerts))
    crit_al    = sum(1 for a in reg_alerts if (a.get("severity") or "").lower() == "critical")
    total_hosp = max(1, len(reg_hosp))
    danger_h   = sum(1 for h in reg_hosp if (h.get("hospital_status") or "").lower() == "dangerous")
    total_rd   = max(MIN_D, len(reg_roads))
    closed_rd  = sum(1 for r in reg_roads if (r.get("status") or r.get("road_type") or "").lower() in ("closed","blocked"))

    score = (
        (active_inc / total_inc) * 40 +
        (crit_al    / total_al)  * 30 +
        (danger_h   / max(1, total_hosp)) * 20 +
        (closed_rd  / total_rd)  * 10
    )
    return min(100, round(score * 100))



# SHEET 1 — SUMMARY DASHBOARD 

def sheet_summary(wb, d):
    ws = wb.create_sheet("Summary Dashboard")
    ws.sheet_view.showGridLines = False
    NC = 8

    subtitle = f"Report Period: {d['dateFrom']} → {d['dateTo']}  |  Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    add_title(ws, "Crisis Management — Executive Summary Dashboard", subtitle, NC)

    incidents = d["incidents"]
    alerts    = d["alerts"]
    hospitals = d["hospitals"]

    csi = crisis_severity_index(d)
    if csi >= 71: csi_fg = C_RED
    elif csi >= 31: csi_fg = C_AMBER
    else: csi_fg = C_GREEN

    ws.row_dimensions[3].height = 8
    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 30
    ws.row_dimensions[6].height = 8

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=NC)
    csi_lbl = ws.cell(row=4, column=1, value="▶  CRISIS SEVERITY INDEX  —  matches the live dashboard badge in Reports & Analytics")
    csi_lbl.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
    csi_lbl.fill = _fill(NAVY); csi_lbl.alignment = _left()

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=NC)
    csi_cell = ws.cell(row=5, column=1, value=f"CSI Score: {csi} / 100")
    csi_cell.font = Font(name="Calibri", bold=True, color=csi_fg, size=14)
    csi_cell.fill = _fill(SILVER); csi_cell.alignment = _center(); csi_cell.border = _border()

    ws.row_dimensions[7].height = 10

    # KPI Cards 
    kpis = [
        ("Total Incidents",  len(incidents)),
        ("Active Incidents", sum(1 for r in incidents if r.get("status","").lower() != "resolved")),
        ("Resolved",         sum(1 for r in incidents if r.get("status","").lower() == "resolved")),
        ("Total Alerts",     len(alerts)),
        ("Critical Alerts",  sum(1 for a in alerts if (a.get("severity") or "").lower() == "critical")),
        ("Hospitals",        len(hospitals)),
        ("Safe Hospitals",   sum(1 for h in hospitals if (h.get("hospital_status") or "").lower() == "safe")),
        ("Dangerous Hosps",  sum(1 for h in hospitals if (h.get("hospital_status") or "").lower() == "dangerous")),
    ]
    ws.row_dimensions[8].height = 26
    ws.row_dimensions[9].height = 52
    for i, (label, value) in enumerate(kpis):
        col = i + 1
        lbg, vcol = KPI_PAIRS[i]
        lc = ws.cell(row=8, column=col, value=label)
        lc.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
        lc.fill = _fill(lbg); lc.alignment = _center(); lc.border = _border()
        vc = ws.cell(row=9, column=col, value=value)
        vc.font = Font(name="Calibri", bold=True, color=lbg, size=24)
        vc.fill = _fill(SILVER); vc.alignment = _center(); vc.border = _border()

    ws.row_dimensions[10].height = 12

    # Response Time Summary
    r = 11
    add_section(ws, r, "⏱   Response Time Summary", NC, NAVY_MID)

    ws.row_dimensions[r+1].height = 24
    ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=7)
    hc = ws.cell(row=r+1, column=1, value="Metric")
    hc.font = Font(name="Calibri", bold=True, color=NAVY, size=10)
    hc.fill = _fill(GRAY_HDR); hc.alignment = _center(); hc.border = _border()
    vc = ws.cell(row=r+1, column=8, value="Value")
    vc.font = Font(name="Calibri", bold=True, color=NAVY, size=10)
    vc.fill = _fill(GRAY_HDR); vc.alignment = _center(); vc.border = _border()

    resolved  = [x for x in incidents if x.get("response_minutes") is not None]
    resp_vals = [float(x["response_minutes"]) for x in resolved]
    pct_with_resp = f"{round(len(resolved)/max(1,len(incidents))*100)}%"

    rt_data = [
        ("Incidents with Recorded Response Time", f"{len(resolved)} ({pct_with_resp})"),
        ("Average Response Time",                 hhmm(safe_avg(resp_vals))),
        ("Fastest Response",                      hhmm(safe_min(resp_vals))),
        ("Slowest Response",                      hhmm(safe_max(resp_vals))),
    ]
    for i, (label, val) in enumerate(rt_data):
        rr = r + 2 + i
        ws.row_dimensions[rr].height = 24
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=7)
        bg = STRIPE_A if i % 2 else STRIPE_B
        lc = ws.cell(row=rr, column=1, value=label)
        lc.font = _font(color="2D3748"); lc.fill = _fill(bg)
        lc.alignment = _left(); lc.border = _border()
        vc2 = ws.cell(row=rr, column=8, value=val)
        vc2.font = Font(name="Calibri", bold=True, color=NAVY, size=12)
        vc2.fill = _fill(bg); vc2.alignment = _center(); vc2.border = _border()

    ws.row_dimensions[r + 6].height = 12

# SHEET 2 — INCIDENTS

def sheet_incidents(wb, d):
    ws = wb.create_sheet("Incidents & Response Time")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"
    NC = 9

    add_title(ws, "Incidents & Response Time Analysis",
              f"Period: {d['dateFrom']} to {d['dateTo']}", NC)

    incidents = d["incidents"]
    resolved  = [x for x in incidents if x.get("response_minutes") is not None]
    resp_vals = [float(x["response_minutes"]) for x in resolved]

    stats = [
        ("Total",        len(incidents),              NAVY_MID),
        ("Active",       sum(1 for r in incidents if r.get("status","").lower() not in ("resolved",)), C_RED),
        ("Resolved",     len(resolved),               C_GREEN),
        ("Avg Response", hhmm(safe_avg(resp_vals)),   C_AMBER),
        ("Fastest",      hhmm(safe_min(resp_vals)),   C_GREEN),
        ("Slowest",      hhmm(safe_max(resp_vals)),   C_RED),
    ]
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 36
    for i, (label, val, color) in enumerate(stats):
        col = i + 1
        lc = ws.cell(row=3, column=col, value=label)
        lc.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
        lc.fill = _fill(color); lc.alignment = _center(); lc.border = _border()
        vc = ws.cell(row=4, column=col, value=val)
        vc.font = Font(name="Calibri", bold=True, color=color, size=14)
        vc.fill = _fill(SILVER); vc.alignment = _center(); vc.border = _border()

    headers = ["#", "ID", "Title", "Region", "Severity", "Status", "Created At", "Resolved At", "Response Time"]
    add_header(ws, 5, headers, height=26)
    ws.auto_filter.ref = "A5:I5"

    for i, row in enumerate(incidents):
        r = 6 + i
        resp_min = row.get("response_minutes")
        vals = [
            i+1, row.get("id",""), row.get("incident_name",""), row.get("location",""),
            row.get("severity",""), row.get("status",""),
            row.get("reported_at",""), row.get("resolved_at","") or "—", hhmm(resp_min),
        ]
        add_row(ws, r, vals, alt=i%2==1, height=26)
        sv = row.get("severity","").lower()
        color_cell(ws, r, 5, SEV_COLOR.get(sv, NAVY))
        st = row.get("status","").lower()
        stc = C_GREEN if st=="resolved" else (C_AMBER if "progress" in st else C_RED)
        color_cell(ws, r, 6, stc)
        if resp_min is not None:
            rf = float(resp_min)
            color_cell(ws, r, 9, C_GREEN if rf < 30 else (C_AMBER if rf < 120 else C_RED))

    widths = {"A":5,"B":6,"C":36,"D":18,"E":12,"F":14,"G":22,"H":22,"I":16}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w






# SHEET 3 — HOSPITALS & CASUALTIES
def sheet_hospitals(wb, d):
    ws = wb.create_sheet("Hospitals & Casualties")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    NC = 10

    add_title(ws, "Hospitals, Capacity & Field Teams",
              "Operational snapshot — bed capacity, staffing, and team deployment", NC)

    martyrs_by_hosp = {}
    for dem in d.get("demographics", []):
        hname = dem.get("hospital_name","")
        total_m = (int(dem.get("male_martyrs",0) or 0) +
                   int(dem.get("female_martyrs",0) or 0) +
                   int(dem.get("children_martyrs",0) or 0))
        martyrs_by_hosp[hname] = martyrs_by_hosp.get(hname, 0) + total_m

    h_headers = ["#", "Hospital Name", "Region", "Status", "Readiness",
                 "Total Beds", "Available Beds", "ICU Beds",
                 "Staff on Duty", "Martyrs"]
    add_header(ws, 3, h_headers, height=26)

    for i, h in enumerate(d["hospitals"]):
        r = 4 + i
        rs = readiness_score(h)
        rl, rc = readiness_label(rs)
        hname = h.get("hospital_name","")
        vals = [
            i+1, hname, h.get("location",""),
            h.get("hospital_status",""), rl,
            h.get("total_beds","") or 0, h.get("available_beds","") or 0,
            h.get("icu_beds","") or 0, h.get("staff_on_duty","") or 0,
            martyrs_by_hosp.get(hname, 0),
        ]
        add_row(ws, r, vals, alt=i%2==1, height=28)
        st = (h.get("hospital_status") or "").lower()
        color_cell(ws, r, 4, ST_HOSP.get(st, C_AMBER))
        color_cell(ws, r, 5, rc)
        if martyrs_by_hosp.get(hname, 0) > 0:
            color_cell(ws, r, 10, C_RED)

    teams_start = 4 + len(d["hospitals"]) + 2
    add_section(ws, teams_start, "🚑   Field Teams & Deployment Status", NC, NAVY_MID)
    t_headers = ["#", "Team Name", "Hospital", "Region", "Status", "Current Location", "Created At", "", "", ""]
    add_header(ws, teams_start+1, t_headers, height=26)

    teams = d.get("hospitalTeams", [])
    for i, t in enumerate(teams):
        rr = teams_start + 2 + i
        vals = [
            i+1, t.get("team_name",""), t.get("hospital_name",""),
            t.get("region",""), t.get("status",""),
            t.get("current_location",""), t.get("created_at",""),
            "", "", "",
        ]
        add_row(ws, rr, vals, alt=i%2==1, height=28)
        ts = (t.get("status") or "").lower()
        color_cell(ws, rr, 5, ST_TEAM.get(ts, C_AMBER))

    demo_start = teams_start + 2 + len(teams) + 2
    add_section(ws, demo_start, "💊   Casualty Demographics by Hospital", NC, C_RED)
    d_headers = ["Hospital", "Male Inj.", "Female Inj.", "Children Inj.",
                 "Male Martyrs", "Female Martyrs", "Children Martyrs",
                 "Total Injured", "Total Martyrs", "Recorded At"]
    add_header(ws, demo_start+1, d_headers, height=26)

    for i, row in enumerate(d["demographics"]):
        rr = demo_start + 2 + i
        mi = int(row.get("male_injured",0) or 0)
        fi = int(row.get("female_injured",0) or 0)
        ci = int(row.get("children_injured",0) or 0)
        mm = int(row.get("male_martyrs",0) or 0)
        fm = int(row.get("female_martyrs",0) or 0)
        cm = int(row.get("children_martyrs",0) or 0)
        vals = [row.get("hospital_name",""), mi, fi, ci, mm, fm, cm,
                mi+fi+ci, mm+fm+cm, row.get("recorded_at","")]
        add_row(ws, rr, vals, alt=i%2==1, height=28)
        if mi+fi+ci > 0: color_cell(ws, rr, 8, C_AMBER)
        if mm+fm+cm > 0: color_cell(ws, rr, 9, C_RED)

    widths = {"A":5,"B":28,"C":20,"D":14,"E":14,"F":12,"G":16,"H":12,"I":14,"J":22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# SHEET 4 — SHELTERS  
def sheet_shelters(wb, d):
    ws = wb.create_sheet("Shelters")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    NC = 7

    add_title(ws, "Shelter Occupancy & Needs Fulfillment", "Current shelter data and needs tracking across all regions", NC)
    add_header(ws, 3, ["#", "Shelter Name", "Region", "Status", "Capacity", "Occupied", "Available"], height=26)

    for i, s in enumerate(d["shelters"]):
        r = 4 + i
        cap   = int(s.get("capacity",0) or 0)
        occ   = int(s.get("occupied",0)  or 0)
        avail = int(s.get("available", cap-occ) or 0)
        vals  = [i+1, s.get("name",""), s.get("region",""), s.get("status",""), cap, occ, avail]
        add_row(ws, r, vals, alt=i%2==1, height=28)
        pct = (occ/cap*100) if cap else 0
        color_cell(ws, r, 6, C_RED if pct > 85 else (C_AMBER if pct > 60 else C_GREEN))

    #  Needs Fulfillment Summary 
    needs_start = 4 + len(d["shelters"]) + 2
    add_section(ws, needs_start, "📦   Needs Fulfillment Summary by Category", NC, NAVY_MID)
    add_header(ws, needs_start+1, ["Category", "Fulfilled", "In Progress", "Not Fulfilled", "Total", "% Fulfilled", ""], height=26)

    needs_by_cat = defaultdict(lambda: {"f":0,"ip":0,"o":0})
    for n in d.get("needs", []):
        cat = (n.get("category") or "Other").capitalize()
        st  = (n.get("status") or "").lower()
        if st == "fulfilled":    needs_by_cat[cat]["f"]  += 1
        elif "progress" in st:   needs_by_cat[cat]["ip"] += 1
        else:                    needs_by_cat[cat]["o"]  += 1

    for i, (cat, vals) in enumerate(sorted(needs_by_cat.items())):
        rr = needs_start + 2 + i
        total = vals["f"] + vals["ip"] + vals["o"]
        pct_f = f"{round(vals['f']/total*100)}%" if total else "0%"
        ws.merge_cells(start_row=rr, start_column=7, end_row=rr, end_column=7)
        add_row(ws, rr, [cat, vals["f"], vals["ip"], vals["o"], total, pct_f, ""], alt=i%2==1)
        color_cell(ws, rr, 2, C_GREEN)
        color_cell(ws, rr, 4, C_RED if vals["o"] > 0 else C_GREEN)
        color_cell(ws, rr, 6, C_GREEN if vals["f"] == total else (C_AMBER if vals["o"] == 0 else C_RED))

    widths = {"A":18,"B":32,"C":18,"D":14,"E":12,"F":12,"G":12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# SHEET 5 — ROADS & ROUTES 

def sheet_roads(wb, d):
    ws = wb.create_sheet("Roads & Routes")
    ws.sheet_view.showGridLines = False
    NC = 5

    add_title(ws, "Roads & Routes — Summary", "Affected road counts reported by police and municipality", NC)

    all_roads = d.get("mapRoads", []) + d.get("polRoads", [])

    # Count by status
    total      = len(all_roads)
    open_count = sum(1 for r in all_roads if (r.get("status") or r.get("road_type") or "").lower() in ("open", "safe"))
    warn_count = sum(1 for r in all_roads if (r.get("status") or r.get("road_type") or "").lower() in ("warning", "restricted"))
    blocked    = sum(1 for r in all_roads if (r.get("status") or r.get("road_type") or "").lower() in ("closed", "blocked", "dangerous", "danger"))
    other      = total - open_count - warn_count - blocked

    routes     = d.get("mapRoutes", [])
    total_routes = len(routes)

    # KPI row 
    kpi_data = [
        ("Total Roads",        total,        NAVY_MID),
        ("Open / Safe",        open_count,   C_GREEN),
        ("Warning / Restricted", warn_count, C_AMBER),
        ("Closed / Blocked",   blocked,      C_RED),
        ("Total Routes",       total_routes, C_TEAL),
    ]

    ws.row_dimensions[3].height = 26
    ws.row_dimensions[4].height = 52
    for i, (label, value, color) in enumerate(kpi_data):
        col = i + 1
        lc = ws.cell(row=3, column=col, value=label)
        lc.font = Font(name="Calibri", bold=True, color=WHITE, size=9)
        lc.fill = _fill(color); lc.alignment = _center(); lc.border = _border()
        vc = ws.cell(row=4, column=col, value=value)
        vc.font = Font(name="Calibri", bold=True, color=color, size=28)
        vc.fill = _fill(SILVER); vc.alignment = _center(); vc.border = _border()

    ws.row_dimensions[5].height = 12

    # Roads breakdown table 
    r = 6
    add_section(ws, r, "🛣️   Road Status Breakdown", NC, NAVY_MID)
    add_header(ws, r+1, ["Status", "Count", "% of Total", "Source", ""], height=26)

    breakdown = [
        ("Open / Safe",          open_count, "map_roads + police_roads", C_GREEN),
        ("Warning / Restricted", warn_count, "map_roads + police_roads", C_AMBER),
        ("Closed / Blocked",     blocked,    "map_roads + police_roads", C_RED),
    ]
    for i, (label, cnt, source, color) in enumerate(breakdown):
        rr = r + 2 + i
        pct = f"{round(cnt / max(1, total) * 100)}%" if total else "0%"
        add_row(ws, rr, [label, cnt, pct, source, ""], alt=i%2==1, height=28)
        color_cell(ws, rr, 1, color)
        color_cell(ws, rr, 2, color)

    # totals row
    tot_row = r + 2 + len(breakdown)
    ws.row_dimensions[tot_row].height = 28
    for col_idx, val in enumerate(["TOTAL", total, "100%", "—", ""], start=1):
        c = ws.cell(row=tot_row, column=col_idx, value=val)
        c.font = Font(name="Calibri", bold=True, color=NAVY, size=10)
        c.fill = _fill(GRAY_HDR); c.alignment = _center(); c.border = _border()

    ws.row_dimensions[tot_row + 1].height = 12

    widths = {"A": 24, "B": 14, "C": 14, "D": 28, "E": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# SHEET 6 — MAP ZONES

def sheet_zones(wb, d):
    ws = wb.create_sheet("Map Zones")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    NC = 6

    add_title(ws, "Map Zones — Safety Overview", "Active zones by region and type", NC)
    add_header(ws, 3, ["#", "Zone Name", "Region", "Type", "Radius (m)", "Created At"], height=26)

    zones = d.get("mapZones", [])
    for i, z in enumerate(zones):
        r = 4 + i
        vals = [i+1, z.get("name",""), z.get("region","") or "—",
                z.get("type",""), z.get("radius_meters",""), z.get("created_at","")]
        add_row(ws, r, vals, alt=i%2==1, height=28)
        color_cell(ws, r, 4, ST_ZONE.get((z.get("type") or "").lower(), C_AMBER))

    sum_start = 4 + len(zones) + 2
    add_section(ws, sum_start, "📊   Zone Summary by Region", NC, NAVY_MID)
    add_header(ws, sum_start+1, ["Region", "Safe Zones", "Warning Zones", "Danger Zones", "Total", ""], height=26)

    by_region = defaultdict(lambda: {"safe":0,"warning":0,"danger":0})
    for z in zones:
        reg = z.get("region","Unknown") or "Unknown"
        zt  = (z.get("type") or "").lower()
        if zt == "safe":      by_region[reg]["safe"]    += 1
        elif zt == "warning": by_region[reg]["warning"] += 1
        else:                 by_region[reg]["danger"]  += 1

    for i, (reg, counts) in enumerate(sorted(by_region.items())):
        rr = sum_start + 2 + i
        total = counts["safe"] + counts["warning"] + counts["danger"]
        add_row(ws, rr, [reg, counts["safe"], counts["warning"], counts["danger"], total, ""], alt=i%2==1, height=28)
        color_cell(ws, rr, 2, C_GREEN)
        color_cell(ws, rr, 3, C_AMBER if counts["warning"] > 0 else C_GREEN)
        color_cell(ws, rr, 4, C_RED if counts["danger"] > 0 else C_GREEN)

    widths = {"A":5,"B":34,"C":16,"D":14,"E":14,"F":22}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# SHEET 7 — ALERTS

def sheet_alerts(wb, d):
    ws = wb.create_sheet("Alerts")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"
    NC = 5

    add_title(ws, "Alerts Log", f"Period: {d['dateFrom']} to {d['dateTo']}", NC)
    add_header(ws, 3, ["#", "Title", "Region", "Severity", "Created At"], height=26)
    ws.auto_filter.ref = "A3:E3"

    for i, a in enumerate(d["alerts"]):
        r = 4 + i
        vals = [i+1, a.get("alert_message",""), a.get("region",""),
                a.get("severity",""), a.get("created_at","")]
        add_row(ws, r, vals, alt=i%2==1, height=28)
        color_cell(ws, r, 4, SEV_COLOR.get((a.get("severity") or "").lower(), C_AMBER))

    widths = {"A":5,"B":50,"C":18,"D":14,"E":24}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def predict_next(vals, window=3):
    """
    Weighted 3-month moving average + linear trend.
    Predicts the next value in a monthly series.
    """
    if not vals:
        return 0
    if len(vals) == 1:
        return round(vals[0])

    recent = vals[-window:] if len(vals) >= window else vals
    base = weighted_avg(recent)

    # linear trend from the last two points
    trend = vals[-1] - vals[-2]

    predicted = base + trend
    return max(0, round(predicted))
# SHEET 8 — TREND & PREDICTIVE ANALYSIS

def sheet_trends(wb, d):
    ws = wb.create_sheet("Trend & Predictive Analysis")
    ws.sheet_view.showGridLines = False
    NC = 7

    add_title(ws, "Trend & Predictive Analysis",
              "Six-month trend analysis, weighted projections & strategic recommendations", NC)

    monthly   = d["monthlyIncidents"]
    regions   = d["regionStats"]
    sh_trend  = d["shelterTrend"]
    hospitals = d.get("hospitals", [])
    teams     = d.get("hospitalTeams", [])
    alerts_d  = d.get("alerts", [])
    roads_d   = d.get("mapRoads", []) + d.get("polRoads", [])
    shelters  = d.get("shelters", [])

    #  Monthly Incident Trend
    r = 3
    add_section(ws, r, "📈   Monthly Incident Trend (Last 6 Months)", NC, NAVY_MID)
    add_header(ws, r+1, ["Month","Total","Resolved","Active","Avg Response Time","MoM Change","Trend"], height=26)

    for i, row in enumerate(monthly):
        rr = r + 2 + i
        prior_total = monthly[i-1]["total"] if i > 0 else None
        change = ""
        if prior_total is not None:
            curr = int(row.get("total",0) or 0); prev = int(prior_total or 0)
            if prev > 0:
                pct = ((curr - prev) / prev) * 100
                change = f"{'▲' if pct>0 else '▼'} {abs(pct):.1f}%"
        trend_arrow = "→ Stable"
        if i > 0:
            curr_val = int(row.get("total",0) or 0)
            prev_val = int(monthly[i-1].get("total",0) or 0)
            trend_arrow = "▲ Rising" if curr_val > prev_val else ("▼ Falling" if curr_val < prev_val else "→ Stable")
        vals = [row.get("month",""), int(row.get("total",0) or 0),
                int(row.get("resolved",0) or 0), int(row.get("active",0) or 0),
                hhmm(row.get("avg_response_min")), change, trend_arrow]
        add_row(ws, rr, vals, alt=i%2==1, height=28)
        if "▲" in str(change): color_cell(ws, rr, 6, C_RED)
        elif "▼" in str(change): color_cell(ws, rr, 6, C_GREEN)

    pred_row = r + 2 + len(monthly) + 1
    if len(monthly) >= 2:
        vals_hist = [int(m.get("total",0) or 0) for m in monthly]
        predicted = predict_next(vals_hist)
        ws.merge_cells(start_row=pred_row, start_column=1, end_row=pred_row, end_column=NC)
        ws.row_dimensions[pred_row].height = 26
        pc = ws.cell(row=pred_row, column=1,
            value=f"📊  Projection (Weighted 3-Month Moving Average + Linear Trend): "
                  f"Estimated {predicted} incidents next month.")
        pc.font = Font(name="Calibri", bold=True, color=C_PURPLE, size=9, italic=True)
        pc.fill = _fill("F3F0FF"); pc.alignment = _left(); pc.border = _border()

    # Regional Activity Score
    r2 = pred_row + 2
    add_section(ws, r2, "🗺️   Regional Activity Score  (≠ global CSI — see note below)", NC, NAVY_MID)
    add_header(ws, r2+1,
               ["Rank","Region","Active Incidents","Critical Alerts","Dangerous Hosp","Blocked Roads","Activity Score (0-100)"],
               height=26)

    scored_regions = []
    for row in regions:
        reg = row.get("region","")
        ras = region_activity_score(reg, d)
        scored_regions.append((ras, row))
    scored_regions.sort(key=lambda x: x[0], reverse=True)

    for i, (ras, row) in enumerate(scored_regions):
        rr = r2 + 2 + i
        reg = row.get("region","")
        reg_alerts   = sum(1 for a in alerts_d if (a.get("region") or "").lower() == reg.lower() and (a.get("severity") or "").lower() == "critical")
        reg_danger_h = sum(1 for h in hospitals if (h.get("location") or "").lower() == reg.lower() and (h.get("hospital_status") or "").lower() == "dangerous")
        reg_closed_r = sum(1 for r_ in roads_d if (r_.get("region") or "").lower() == reg.lower() and (r_.get("status") or r_.get("road_type") or "").lower() in ("closed","blocked"))

        risk_lbl = "HIGH" if ras >= 60 else ("MEDIUM" if ras >= 30 else "LOW")
        vals = [i+1, reg, int(row.get("active",0) or 0), reg_alerts, reg_danger_h, reg_closed_r, f"{ras}/100  [{risk_lbl}]"]
        add_row(ws, rr, vals, alt=i%2==1, height=28)
        color_cell(ws, rr, 7, C_RED if ras >= 60 else (C_AMBER if ras >= 30 else C_GREEN))

    # Note row explaining the score
    note_r = r2 + 2 + len(scored_regions)
    ws.row_dimensions[note_r].height = 36
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=NC)
    nc = ws.cell(row=note_r, column=1,
        value="ℹ️  Activity Score formula: 40% Active Incidents + 30% Critical Alerts + 20% Dangerous Hospitals + 10% Blocked Roads. "
              "Uses a minimum denominator of 5 per factor — regions with only 1-2 incidents will not score 100%. "
              "This is a PER-REGION activity index and is NOT the same as the global Crisis Severity Index on the Summary sheet.")
    nc.font = Font(name="Calibri", italic=True, color=GRAY_TXT, size=8)
    nc.fill = _fill(SILVER); nc.alignment = _left(indent=1); nc.border = _border()

    # Hospital Infrastructure Analysis 
    r3 = note_r + 2
    add_section(ws, r3, "🏥   Hospital Infrastructure & Operational Readiness", NC, NAVY_MID)
    add_header(ws, r3+1, ["Hospital Name", "Region", "Status", "Infrastructure", "Power", "Water", "Readiness"], height=26)

    hosp_total = len(hospitals)
    infra_issues = power_issues = water_issues = 0
    hosp_safe = hosp_warn = hosp_danger = 0

    for i, h in enumerate(hospitals):
        rr = r3 + 2 + i
        h_status = (h.get("hospital_status") or "").lower()
        infra    = (h.get("infrastructure_status") or "").lower()
        power    = (h.get("power_status") or "").lower()
        water    = (h.get("water_status") or "").lower()

        if h_status == "safe":      hosp_safe   += 1
        elif h_status == "warning": hosp_warn   += 1
        else:                       hosp_danger += 1
        if infra not in ("intact",):            infra_issues += 1
        if power not in ("stable","available"): power_issues += 1
        if water not in ("stable","available"): water_issues += 1

        rs = readiness_score(h)
        rl, rc = readiness_label(rs)
        vals = [h.get("hospital_name",""), h.get("location",""),
                h.get("hospital_status",""), h.get("infrastructure_status",""),
                h.get("power_status",""), h.get("water_status",""), f"{rl}  ({rs}/100)"]
        add_row(ws, rr, vals, alt=i%2==1, height=28)
        color_cell(ws, rr, 3, ST_HOSP.get(h_status, C_AMBER))
        color_cell(ws, rr, 4, ST_INFRA.get(infra, C_AMBER))
        color_cell(ws, rr, 5, ST_UTIL.get(power, C_AMBER))
        color_cell(ws, rr, 6, ST_UTIL.get(water, C_AMBER))
        color_cell(ws, rr, 7, rc)

    # Hospital Capacity Recommendations 
    hosp_cap_start = r3 + 2 + hosp_total + 1
    add_section(ws, hosp_cap_start, "🏥   Hospital Capacity Recommendations", NC, NAVY_LIGHT)
    add_header(ws, hosp_cap_start+1, ["Hospital Name", "Region", "Readiness", "Recommendation", "", "", ""], height=26)

    for i, h in enumerate(hospitals):
        rr = hosp_cap_start + 2 + i
        hname  = h.get("hospital_name","")
        region = h.get("location","")
        rs     = readiness_score(h)
        rl, rc = readiness_label(rs)
        reg_inc = sum(1 for inc in d.get("incidents",[]) if (inc.get("location") or "").lower() == region.lower())
        reg_danger_zones = sum(1 for z in d.get("mapZones",[]) if (z.get("region") or "").lower() == region.lower() and (z.get("type") or "").lower() == "danger")
        ambulances = int(h.get("ambulances",0) or 0)
        avail_beds = int(h.get("available_beds",0) or 0)
        icu_avail  = int(h.get("available_icu_beds",0) or 0)

        if rl == "Critical":
            rec = (f"{hname}: CRITICAL ({rs}/100). Region has {reg_inc} incidents, {reg_danger_zones} danger zone(s). "
                   f"Immediate: coordinate patient transfers, deploy emergency support teams, restore infrastructure.")
            rec_color = C_RED
        elif rl == "Medium":
            rec = (f"{hname}: Medium readiness ({rs}/100), {ambulances} ambulance(s). "
                   f"Region: {reg_inc} incident(s), {reg_danger_zones} danger zone(s). "
                   f"Pre-position additional medical resources and monitor capacity trends.")
            rec_color = C_AMBER
        else:
            rec = (f"{hname}: High readiness ({rs}/100). "
                   f"Available beds: {avail_beds}, ICU: {icu_avail}, Ambulances: {ambulances}. Maintain standard operations.")
            rec_color = C_GREEN

        ws.row_dimensions[rr].height = 60
        bg = STRIPE_A if i%2==1 else STRIPE_B
        for col_idx, val in enumerate([hname, region, f"{rl} ({rs}/100)"], start=1):
            c = ws.cell(row=rr, column=col_idx, value=val)
            c.font = _font(color="2D3748"); c.fill = _fill(bg)
            c.alignment = _left(); c.border = _border()
        color_cell(ws, rr, 3, rc)
        ws.merge_cells(start_row=rr, start_column=4, end_row=rr, end_column=NC)
        rec_cell = ws.cell(row=rr, column=4, value=rec)
        rec_cell.font = Font(name="Calibri", color=rec_color, size=9)
        rec_cell.fill = _fill(bg)
        rec_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        rec_cell.border = _border()

    # Resource Gap Analysis 
    gap_start = hosp_cap_start + 2 + len(hospitals) + 1
    add_section(ws, gap_start, "📊   Resource Gap Analysis", NC, C_TEAL)
    add_header(ws, gap_start+1, ["Resource", "Estimated Needed", "Available", "Gap", "Status", "Priority", ""], height=26)

    total_amb_available = sum(int(h.get("ambulances",0) or 0) for h in hospitals)
    total_icu_available = sum(int(h.get("available_icu_beds",0) or 0) for h in hospitals)
    total_beds_avail    = sum(int(h.get("available_beds",0) or 0) for h in hospitals)
    cap_total   = sum(int(s.get("capacity",0) or 0) for s in shelters)
    occ_total   = sum(int(s.get("occupied",0)  or 0) for s in shelters)
    sh_available = cap_total - occ_total
    active_count = sum(1 for inc in d.get("incidents",[]) if (inc.get("status") or "").lower() not in ("resolved",))
    available_teams = sum(1 for t in teams if (t.get("status") or "").lower() == "available")

    amb_needed  = max(active_count * 2, 10)
    icu_needed  = max(active_count * 3, 20)
    beds_needed = max(active_count * 5, 50)
    shelter_needed = max(occ_total + round(occ_total * 0.15), occ_total)

    gap_rows = [
        ("Ambulances",     amb_needed,  total_amb_available, amb_needed  - total_amb_available),
        ("ICU Beds",       icu_needed,  total_icu_available, icu_needed  - total_icu_available),
        ("Hospital Beds",  beds_needed, total_beds_avail,    beds_needed - total_beds_avail),
        ("Shelter Capacity", shelter_needed, sh_available,   shelter_needed - sh_available),
        ("Field Teams",    max(active_count, len(teams)), available_teams,
                           max(active_count, len(teams)) - available_teams),
    ]

    for i, (resource, needed, avail, gap) in enumerate(gap_rows):
        rr = gap_start + 2 + i
        gap_str  = f"−{max(0,gap)}" if gap > 0 else (f"+{abs(gap)} surplus" if gap < 0 else "Met")
        status   = "DEFICIT" if gap > 0 else ("SURPLUS" if gap < 0 else "BALANCED")
        priority = "URGENT" if (gap > 0 and i <= 1) else ("HIGH" if gap > 0 else "LOW")
        vals = [resource, needed, avail, gap_str, status, priority, ""]
        add_row(ws, rr, vals, alt=i%2==1, height=28)
        color_cell(ws, rr, 4, C_RED if gap > 0 else C_GREEN)
        color_cell(ws, rr, 5, C_RED if status == "DEFICIT" else (C_GREEN if status == "SURPLUS" else C_AMBER))
        color_cell(ws, rr, 6, C_RED if priority == "URGENT" else (C_AMBER if priority == "HIGH" else C_GREEN))

    gap_note_row = gap_start + 2 + len(gap_rows)
    ws.merge_cells(start_row=gap_note_row, start_column=1, end_row=gap_note_row, end_column=NC)
    ws.row_dimensions[gap_note_row].height = 22
    gn = ws.cell(row=gap_note_row, column=1,
        value="Note: Required values estimated from active incident count and operational standards. Actual requirements may vary.")
    gn.font = Font(name="Calibri", italic=True, color=GRAY_TXT, size=8)
    gn.fill = _fill(SILVER); gn.alignment = _left(); gn.border = _border()

    #Shelter Occupancy Trend 
    r4 = gap_note_row + 2
    add_section(ws, r4, "🏠   Shelter Occupancy Trend (Last 6 Months)", NC, NAVY_MID)
    add_header(ws, r4+1, ["Month","Avg Capacity","Avg Occupied","Occupancy %","Availability %","MoM Change",""], height=26)

    for i, row in enumerate(sh_trend):
        rr = r4 + 2 + i
        cap = float(row.get("avg_capacity",0) or 0)
        occ = float(row.get("avg_occupied",0)  or 0)
        pct_occ   = f"{(occ/cap*100):.1f}%" if cap else "0.0%"
        pct_avail = f"{((cap-occ)/cap*100):.1f}%" if cap else "0.0%"
        mom = ""
        if i > 0:
            prev_occ = float(sh_trend[i-1].get("avg_occupied",0) or 0)
            diff = occ - prev_occ
            mom = f"{'▲' if diff>0 else '▼'} {abs(diff):.0f}" if diff != 0 else "→"
        vals = [row.get("month",""), round(cap,1), round(occ,1), pct_occ, pct_avail, mom, ""]
        add_row(ws, rr, vals, alt=i%2==1, height=28)
        pv = (occ/cap*100) if cap else 0
        color_cell(ws, rr, 4, C_RED if pv > 85 else (C_AMBER if pv > 60 else C_GREEN))

    pred_r4_base = r4 + 2 + len(sh_trend)
    if len(sh_trend) >= 2:
        occs   = [float(s.get("avg_occupied",0) or 0) for s in sh_trend]
        avg_cap = safe_avg([float(s.get("avg_capacity",0) or 0) for s in sh_trend]) or 1
        pred_occ = predict_next(occs)
        pred_pct = (pred_occ / avg_cap * 100) if avg_cap > 0 else 0
        pred_r4  = pred_r4_base + 1
        ws.merge_cells(start_row=pred_r4, start_column=1, end_row=pred_r4, end_column=NC)
        ws.row_dimensions[pred_r4].height = 26
        prd = ws.cell(row=pred_r4, column=1,
            value=f"📊  Shelter Projection: Estimated {pred_occ:.0f} persons ({pred_pct:.1f}% of capacity) next month.")
        prd.font = Font(name="Calibri", bold=True, color=C_PURPLE, size=9, italic=True)
        prd.fill = _fill("F3F0FF"); prd.alignment = _left(); prd.border = _border()
        analysis_start = pred_r4 + 2
    else:
        analysis_start = pred_r4_base + 2

    # Regional Resource Recommendation 
    rrr_start = analysis_start
    add_section(ws, rrr_start, "📍   Regional Resource Recommendation", NC, NAVY_LIGHT)
    add_header(ws, rrr_start+1,
               ["Region", "Incidents", "Danger Zones", "Hospitals", "Activity Score", "Teams Avail.", "Recommendation"],
               height=26)

    region_names = list({r.get("region","") for r in regions if r.get("region")})
    for i, reg in enumerate(region_names):
        rr = rrr_start + 2 + i
        reg_inc_total = sum(1 for inc in d.get("incidents",[]) if (inc.get("location") or "").lower() == reg.lower())
        reg_danger_z  = sum(1 for z in d.get("mapZones",[]) if (z.get("region") or "").lower() == reg.lower() and (z.get("type") or "").lower() == "danger")
        reg_hosps     = [h for h in hospitals if (h.get("location") or "").lower() == reg.lower()]
        reg_teams_av  = sum(1 for t in teams if (t.get("region") or "").lower() == reg.lower() and (t.get("status") or "").lower() == "available")
        ras = region_activity_score(reg, d)

        if ras >= 60:
            rec = f"HIGH — Deploy additional ambulance teams, increase hospital liaison. {reg_danger_z} danger zone(s) require priority containment."
        elif ras >= 30:
            rec = f"MODERATE — Monitor closely. Pre-position resources near active clusters. Ensure {len(reg_hosps)} hospital(s) maintain staffing."
        else:
            rec = f"LOW — Current resources sufficient. Maintain standard operational posture."

        add_row(ws, rr, [reg, reg_inc_total, reg_danger_z, len(reg_hosps), f"{ras}/100", reg_teams_av, rec],
                alt=i%2==1, height=52)
        color_cell(ws, rr, 5, C_RED if ras >= 60 else (C_AMBER if ras >= 30 else C_GREEN))
        color_cell(ws, rr, 7, C_RED if ras >= 60 else (C_AMBER if ras >= 30 else C_GREEN), bold=False, size=9)

    # Situation Analysis 
    ra = rrr_start + 2 + len(region_names) + 2
    add_section(ws, ra, "📋   Situation Analysis & Strategic Recommendations", NC, NAVY)
    ws.row_dimensions[ra].height = 32

    lines = []
    if len(monthly) >= 2:
        vals_hist2 = [int(m.get("total",0) or 0) for m in monthly]
        predicted2 = predict_next(vals_hist2)
        inc_trend2 = vals_hist2[-1] - vals_hist2[-2]
        last_month2 = monthly[-1].get("month","")
        if inc_trend2 > 0:
            lines.append(("⚠️  Incident Trend",
                f"Incident count increased by {inc_trend2} in {last_month2}. "
                f"Weighted model projects {predicted2} incidents next month. Warrants increased response allocation.",
                C_RED))
        else:
            lines.append(("✅  Incident Trend",
                f"Incident count decreased by {abs(inc_trend2)} in {last_month2}. "
                f"Weighted model projects {predicted2} incidents next month — de-escalation trend.",
                C_GREEN))

    if hosp_danger > 0:
        lines.append(("🚨  Hospital Status",
            f"{hosp_danger} hospital(s) DANGEROUS. Activate overflow protocols immediately and coordinate patient transfers.",
            C_RED))
    elif hosp_warn > 0:
        lines.append(("⚠️  Hospital Status",
            f"{hosp_warn} hospital(s) under WARNING. Monitor closely and consider pre-emptive patient redistribution.",
            C_AMBER))
    else:
        lines.append(("✅  Hospital Status",
            f"All {hosp_safe} hospitals Safe. Maintain regular reporting and adequate supply stockpiles.",
            C_GREEN))

    if infra_issues > 0:
        lines.append(("🚨  Infrastructure",
            f"{infra_issues} hospital(s) report infrastructure damage. Prioritize structural assessment and repair.",
            C_RED))
    if power_issues > 0:
        lines.append(("⚠️  Power Supply",
            f"{power_issues} hospital(s) have unstable power. Deploy emergency generators immediately.",
            C_AMBER))

    all_resp = [float(x.get("response_minutes",0) or 0) for x in d.get("incidents",[]) if x.get("response_minutes")]
    if all_resp:
        avg_r2 = sum(all_resp)/len(all_resp)
        if avg_r2 > 240:
            lines.append(("⏱  Response Time",
                f"Average response {hhmm(avg_r2)} exceeds 2-hour threshold. Review dispatch and pre-position units in high-risk zones.",
                C_RED))
        elif avg_r2 > 60:
            lines.append(("⚠️  Response Time",
                f"Average response {hhmm(avg_r2)} — acceptable but above optimal. Focus optimization on highest-risk regions.",
                C_AMBER))
        else:
            lines.append(("✅  Response Time",
                f"Average response {hhmm(avg_r2)} within acceptable parameters. Maintain current dispatch protocols.",
                C_GREEN))

    LABEL_BG = {"🚨": "FFF0F0", "⚠️": "FFFBEB", "✅": "F0FDF4", "⏱": "FFFBEB"}
    for i, (title, text, color) in enumerate(lines):
        row_t = ra + 1 + (i * 3)
        row_b = row_t + 1
        ws.row_dimensions[row_t].height = 22
        ws.row_dimensions[row_b].height = 52

        ws.merge_cells(start_row=row_t, start_column=1, end_row=row_t, end_column=NC)
        tc = ws.cell(row=row_t, column=1, value=title)
        emoji_key = next((k for k in LABEL_BG if title.startswith(k)), "⏱")
        tc.font = Font(name="Calibri", bold=True, color=color, size=10)
        tc.fill = _fill(LABEL_BG.get(emoji_key, "FFFFFF"))
        tc.alignment = _left(); tc.border = _border()

        ws.merge_cells(start_row=row_b, start_column=1, end_row=row_b, end_column=NC)
        bc = ws.cell(row=row_b, column=1, value=text)
        bc.font = Font(name="Calibri", color="2D3748", size=9)
        bc.fill = _fill(WHITE)
        bc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
        bc.border = _border()

    widths = {"A":22,"B":22,"C":18,"D":18,"E":18,"F":16,"G":42}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

# MAIN
def main():
    if len(sys.argv) < 3:
        print("Usage: generate_report_excel.py <input.json> <output.xlsx>")
        sys.exit(1)

    with open(sys.argv[1], "r", encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)

    sheet_summary(wb, data)
    sheet_incidents(wb, data)
    sheet_hospitals(wb, data)
    sheet_shelters(wb, data)
    sheet_roads(wb, data)
    sheet_zones(wb, data)
    sheet_alerts(wb, data)
    sheet_trends(wb, data)

    wb.save(sys.argv[2])
    print("OK")

if __name__ == "__main__":
    main()